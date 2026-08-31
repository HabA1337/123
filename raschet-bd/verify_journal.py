#!/usr/bin/env python3
"""Проверка книги: форма журнала 660, правило 2 за 3, рапорт с фамилией."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "Журнал_учета_служебного_времени.xlsx"


def rest_days(bd_days: int) -> int:
    return (bd_days // 3) * 2


def compensation(ovd: float, ovz: float, days: int) -> float:
    return (ovd + ovz) / 30.0 * days


def main() -> None:
    assert XLSX.exists(), XLSX
    wb = load_workbook(XLSX)
    expected_sheets = [
        "Как_заполнять",
        "Анкета",
        "Журнал_день",
        "Журнал_660",
        "БД_сутки",
        "Итоги",
        "Рапорт",
    ]
    assert wb.sheetnames == expected_sheets, wb.sheetnames

    ank = wb["Анкета"]
    assert ank["C3"].value == "Петров"
    assert ank["C12"].value == 2026
    assert ank["C13"].value == 25000
    assert ank["C14"].value == 12000

    daily = wb["Журнал_день"]
    headers = [daily.cell(4, c).value for c in range(1, 14)]
    assert "Дата" in headers[1]
    assert "БД" in str(headers[6])
    bd_marks = 0
    for row in range(6, 372):
        if daily.cell(row, 7).value == 1:
            bd_marks += 1
    assert bd_marks == 30, bd_marks

    j660 = wb["Журнал_660"]
    nums = [j660.cell(7, c).value for c in range(1, 14)]
    assert nums == list(range(1, 14)), nums
    title = j660["A2"].value or ""
    assert "ЖУРНАЛ УЧЕТА ВРЕМЕНИ ПРИВЛЕЧЕНИЯ" in title
    assert "30.10.2015" in (j660["A1"].value or "")
    assert "№ 660" in (j660["A1"].value or "")
    # first person pulled from anketa
    assert j660["B8"].value == "=Анкета!C7"
    assert j660["D8"].value == "=Анкета!B19"

    bd = wb["БД_сутки"]
    assert "INT(C4/3)*2" in str(bd["C5"].value).replace(" ", "")
    assert "Журнал_день!G" in str(bd["C4"].value)

    report = wb["Рапорт"]
    assert report["A8"].value == "РАПОРТ"
    body = str(report["A10"].value)
    assert "Анкета!C6" in body and "Анкета!B19" in body
    assert "боевому дежурству" in body
    pay = str(report["A20"].value)
    assert "76-ФЗ" in pay and "14.02.2010 № 80" in pay
    assert "БД_сутки!C10" in pay

    # legal math for demo: 30 BD -> 20 rest days
    assert rest_days(30) == 20
    assert rest_days(2) == 0
    assert rest_days(4) == 2
    money = compensation(25000, 12000, 20)
    assert abs(money - 24666.6667) < 0.01, money

    print("OK sheets", wb.sheetnames)
    print("OK official columns 1-13")
    print("OK demo BD marks", bd_marks)
    print("OK rest for 30 BD =", rest_days(30), "days")
    print("OK compensation", round(money, 2), "RUB")
    print("OK report uses surname from Анкета")


if __name__ == "__main__":
    main()
