#!/usr/bin/env python3
"""Собирает книгу Excel: официальный журнал (приказ МО РФ № 660, прил. 7),
ежедневный учёт, боевое дежурство в сутках и рапорт на компенсацию."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins

OUT = Path(__file__).resolve().parent / "Журнал_учета_служебного_времени.xlsx"

THIN = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
MED = Border(
    left=Side(style="medium", color="000000"),
    right=Side(style="medium", color="000000"),
    top=Side(style="medium", color="000000"),
    bottom=Side(style="medium", color="000000"),
)

FILL_TITLE = PatternFill("solid", fgColor="1F4E3D")
FILL_HEAD = PatternFill("solid", fgColor="2E7D4F")
FILL_SUB = PatternFill("solid", fgColor="DCE8D8")
FILL_IN = PatternFill("solid", fgColor="FFF3B0")
FILL_OUT = PatternFill("solid", fgColor="E8F0FE")
FILL_NOTE = PatternFill("solid", fgColor="F4F1E8")
FILL_WEEKEND = PatternFill("solid", fgColor="F8E0E0")
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")
FILL_TOTAL = PatternFill("solid", fgColor="C8DCC0")

FONT_WHITE = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
FONT_H = Font(name="Times New Roman", size=11, bold=True)
FONT_H9 = Font(name="Times New Roman", size=9, bold=True)
FONT_N = Font(name="Times New Roman", size=11)
FONT_N9 = Font(name="Times New Roman", size=9)
FONT_TITLE = Font(name="Times New Roman", size=14, bold=True, color="FFFFFF")
FONT_SMALL = Font(name="Times New Roman", size=8, italic=True)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
JUST = Alignment(horizontal="justify", vertical="top", wrap_text=True)

DAYS = 366
DAILY_FIRST = 6
DAILY_LAST = DAILY_FIRST + DAYS - 1  # 371
SQUAD_ROWS = 16


def apply_grid(ws, min_row, max_row, min_col, max_col, font=FONT_N, fill=None):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = THIN
            cell.font = font
            cell.alignment = CENTER
            if fill is not None:
                cell.fill = fill


def set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def print_setup(ws, landscape=False, fit=True, paper="A4"):
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = fit
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def build_help(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Как_заполнять"
    ws.sheet_properties.tabColor = "1F4E3D"
    ws.merge_cells("A1:B1")
    ws["A1"] = "Как пользоваться журналом и рапортом"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    lines = [
        (3, "Что это",
         "Книга повторяет официальный журнал из приложения № 7 к приказу Министра обороны РФ "
         "от 30.10.2015 № 660 (п. 25 Порядка) и отдельно ведёт боевое дежурство в сутках "
         "(п. 3 ст. 11 закона № 76-ФЗ и п. 5 приложения № 2 к Положению о порядке прохождения военной службы)."),
        (5, "Жёлтые ячейки",
         "Заполняете только их. Серые / голубые ячейки считаются сами — их не затирайте."),
        (7, "1. Анкета",
         "Впишите фамилию, звание, должность, в/ч, командира, оклады и год. "
         "Фамилия из анкеты сразу попадает в рапорт."),
        (9, "2. Журнал_день — ваш рабочий журнал",
         "Это повседневный учёт на каждый день года. Ставьте 1 в колонке «БД», если стояли "
         "на боевом дежурстве; 1 в «Наряд», если суточный наряд. Сверхурочные часы в рабочие дни — "
         "в графу «Сверхурочно, ч». Часы в выходной или праздник — в графу «Выходной/праздник, ч». "
         "Если дали отгул — часы в «Отдых дан, ч» или сутки в «К отпуску, сут»."),
        (11, "3. Журнал_660 — как в части",
         "Это та самая официальная форма на неделю: графы 1–13, подпись военнослужащего, "
         "подпись командира подразделения. Выберите номер недели — часы подтянутся из Журнал_день. "
         "Можно дописать сослуживцев (журнал в части ведётся на подразделение)."),
        (13, "Важно про журнал 660",
         "Официальный журнал считает ЧАСЫ сверхурочки и выходных (п. 1 ст. 11). "
         "Боевое дежурство в нём отдельной графы не имеет: БД учитывается СУТКАМИ. "
         "Поэтому боевое дежурство смотрите на листе «БД_сутки» и в итогах."),
        (15, "4. БД_сутки",
         "Считает сутки боевого дежурства из ежедневного журнала. "
         "Правило: за каждые 3 суток БД положено 2 суток отдыха. "
         "Остаток (1 или 2 суток) не сгорает — ждёт следующую тройку."),
        (17, "5. Итоги",
         "Сводка за выбранный день, за каждый месяц и за год. "
         "Деньги считаются только за неиспользованные сутки отдыха за БД "
         "по приказу МО РФ № 80: (оклад должности + оклад звания) / 30 × сутки."),
        (19, "6. Рапорт",
         "Собирается сам. Проверьте цифры, распечатайте лист «Рапорт» в двух экземплярах "
         "и сдайте через канцелярию. На своём экземпляре возьмите входящий номер."),
        (21, "Сгорают ли сутки 31 декабря",
         "Нет. Они копятся по мере дежурства и переносятся на следующий год. "
         "Лучше использовать или компенсировать не позже конца следующего года."),
        (23, "Демо-данные",
         "В журнале на 2026 год для примера отмечено 30 суток БД (1–30 января). "
         "Замените на свои отметки: должно получиться 20 суток отдыха."),
    ]
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110
    for row, title, text in lines:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row, 1, title).font = FONT_WHITE
        ws.cell(row, 1).fill = FILL_HEAD
        ws.cell(row, 1).alignment = LEFT
        ws.row_dimensions[row].height = 20
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=2)
        ws.cell(row + 1, 1, text).font = FONT_N
        ws.cell(row + 1, 1).alignment = LEFT
        ws.cell(row + 1, 1).fill = FILL_NOTE
        ws.row_dimensions[row + 1].height = 48
    print_setup(ws, landscape=False)


def build_anketa(wb: Workbook) -> None:
    ws = wb.create_sheet("Анкета")
    ws.sheet_properties.tabColor = "C9A227"
    ws.merge_cells("A1:C1")
    ws["A1"] = "Сведения о военнослужащем — попадают в журнал и в рапорт"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 26

    fields = [
        (3, "Фамилия", "Петров", "C3"),
        (4, "Имя", "Пётр", "C4"),
        (5, "Отчество", "Петрович", "C5"),
        (6, "Воинское звание", "сержант", "C6"),
        (7, "Воинская должность", "командир отделения", "C7"),
        (8, "Подразделение", "1 мсв 2 мср", "C8"),
        (9, "Войсковая часть", "00000", "C9"),
        (10, "Командир части (звание)", "полковник", "C10"),
        (11, "Командир части (инициал и фамилия)", "И.И. Иванов", "C11"),
        (12, "Год учёта", 2026, "C12"),
        (13, "Оклад по воинской должности, руб.", 25000, "C13"),
        (14, "Оклад по воинскому званию, руб.", 12000, "C14"),
        (15, "Продолжительность служебного дня, ч", 8, "C15"),
        (16, "Номер недели для печати журнала 660", 1, "C16"),
    ]
    ws["A2"] = "Реквизит"
    ws["B2"] = "Заполните жёлтое"
    ws["C2"] = "Значение"
    for col in ("A", "B", "C"):
        ws[f"{col}2"].font = FONT_WHITE
        ws[f"{col}2"].fill = FILL_HEAD
        ws[f"{col}2"].alignment = CENTER
        ws[f"{col}2"].border = THIN

    for row, label, value, _ in fields:
        ws.cell(row, 1, label).font = FONT_H
        ws.cell(row, 1).alignment = LEFT
        ws.cell(row, 1).border = THIN
        ws.cell(row, 1).fill = FILL_SUB
        ws.cell(row, 2, "").border = THIN
        cell = ws.cell(row, 3, value)
        cell.fill = FILL_IN
        cell.font = FONT_H
        cell.alignment = CENTER
        cell.border = THIN
        ws.row_dimensions[row].height = 20

    ws["C12"].number_format = "0"
    ws["C13"].number_format = "#,##0"
    ws["C14"].number_format = "#,##0"
    ws["C16"].number_format = "0"

    # Derived full name and addressee
    ws["A18"] = "Служебные поля (не трогать)"
    ws["A18"].font = FONT_H
    ws.merge_cells("A19:A19")
    derived = [
        (19, "ФИО полностью", '=TRIM(C3&" "&C4&" "&C5)'),
        (20, "ФИО в рапорте (род. падеж — правьте при необходимости)", '=C3&"а "&LEFT(C4,1)&"."&LEFT(C5,1)&"."'),
        (21, "Кому рапорт", '="Командиру войсковой части "&C9&" "&C10&"у "&C11'),
        (22, "От кого", '="от "&C7&" "&C6&"а "&C3&"а "&LEFT(C4,1)&"."&LEFT(C5,1)&"."'),
        (23, "Оклады вместе", "=C13+C14"),
        (24, "1/30 окладов", "=C23/30"),
    ]
    for row, label, formula in derived:
        ws.cell(row, 1, label).font = FONT_N9
        ws.cell(row, 1).alignment = LEFT
        ws.cell(row, 1).border = THIN
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        cell = ws.cell(row, 2, formula)
        cell.fill = FILL_OUT
        cell.font = FONT_N
        cell.alignment = LEFT
        cell.border = THIN
        ws.cell(row, 3).border = THIN
        ws.row_dimensions[row].height = 20
    ws["B23"].number_format = "#,##0.00"
    ws["B24"].number_format = "#,##0.00"

    ws["A26"] = (
        "Родительный падеж в строке 20 проверьте сами: «Петрова П.П.» обычно верно, "
        "для фамилий на -о, -ых, -ко поправьте вручную."
    )
    ws.merge_cells("A26:C26")
    ws["A26"].font = FONT_SMALL
    ws["A26"].alignment = LEFT

    ws["A28"] = "Подсказка по окладам"
    ws["A28"].font = FONT_H
    ws.merge_cells("A29:C30")
    ws["A29"] = (
        "Компенсация по приказу МО РФ № 80 считается только из двух окладов "
        "(должность + звание), без надбавок, премии и районного коэффициента. "
        "Берите цифры из расчётного листка."
    )
    ws["A29"].alignment = LEFT
    ws["A29"].fill = FILL_NOTE
    ws["A29"].font = FONT_N

    set_widths(ws, {"A": 62, "B": 28, "C": 36})
    print_setup(ws)
    ws.freeze_panes = "A3"


def build_daily(wb: Workbook) -> None:
    ws = wb.create_sheet("Журнал_день")
    ws.sheet_properties.tabColor = "2E7D4F"
    ws.merge_cells("A1:M1")
    ws["A1"] = (
        "ЕЖЕДНЕВНЫЙ УЧЁТ СЛУЖЕБНОГО ВРЕМЕНИ  —  заполняется на каждого военнослужащего. "
        "Жёлтое — ваши отметки. 1 = заступил. Часы — числом."
    )
    ws["A1"].font = FONT_WHITE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:M2")
    ws["A2"] = (
        '=CONCATENATE("Военнослужащий: ",Анкета!C6," ",Анкета!B19,"   |   ",Анкета!C7,'
        '"   |   в/ч ",Анкета!C9,"   |   год ",Анкета!C12)'
    )
    ws["A2"].font = FONT_H
    ws["A2"].fill = FILL_SUB
    ws["A2"].alignment = CENTER
    ws.row_dimensions[2].height = 20

    ws.merge_cells("A3:M3")
    ws["A3"] = (
        "БД и наряд отмечайте единицей. Сверхурочка журнала 660 — только часы "
        "(графа 5 — рабочие дни, графа 6 — выходные/праздники). Боевое дежурство в часы 660 не переводите: "
        "оно идёт в лист «БД_сутки»."
    )
    ws["A3"].font = FONT_SMALL
    ws["A3"].fill = FILL_NOTE
    ws["A3"].alignment = LEFT

    headers = [
        "№",
        "Дата",
        "День недели",
        "Неделя",
        "Месяц",
        "Вид дня",
        "БД\n(1)",
        "Наряд\n(1)",
        "Сверхурочно,\nч (гр. 5)",
        "Выходной/\nпраздник, ч (гр. 6)",
        "Отдых дан,\nч (гр. 9)",
        "К отпуску,\nсут (гр. 11)",
        "Примечание\n(приказ, график)",
    ]
    for col, title in enumerate(headers, 1):
        cell = ws.cell(4, col, title)
        cell.font = FONT_WHITE
        cell.fill = FILL_HEAD
        cell.alignment = CENTER
        cell.border = THIN
    # numbers row
    for col, num in enumerate(
        ["", "", "", "", "", "", "БД", "наряд", "5", "6", "9", "11", ""], 1
    ):
        cell = ws.cell(5, col, num)
        cell.font = FONT_H9
        cell.fill = FILL_SUB
        cell.alignment = CENTER
        cell.border = THIN
    ws.row_dimensions[4].height = 32
    ws.row_dimensions[5].height = 16
    ws.auto_filter.ref = f"A5:M{DAILY_LAST}"
    ws.freeze_panes = "A6"

    dv01 = DataValidation(type="list", formula1='"1"', allow_blank=True)
    dv01.error = "Ставьте 1 или оставьте пустым"
    dv01.errorTitle = "Отметка"
    ws.add_data_validation(dv01)
    dv01.add(f"G{DAILY_FIRST}:H{DAILY_LAST}")

    for i in range(DAYS):
        row = DAILY_FIRST + i
        n = i + 1
        ws.cell(row, 1, n).font = FONT_N9
        # Date of year; blank after 31 Dec
        ws.cell(row, 2, f'=IF(DATE(Анкета!$C$12,1,1)+{i}>DATE(Анкета!$C$12,12,31),"",DATE(Анкета!$C$12,1,1)+{i})')
        ws.cell(row, 2).number_format = "DD.MM.YYYY"
        ws.cell(
            row,
            3,
            f'=IF(B{row}="","",CHOOSE(WEEKDAY(B{row},2),"понедельник","вторник","среда","четверг","пятница","суббота","воскресенье"))',
        )
        ws.cell(row, 4, f'=IF(B{row}="","",WEEKNUM(B{row},21))')
        ws.cell(row, 5, f'=IF(B{row}="","",MONTH(B{row}))')
        ws.cell(row, 6, f'=IF(B{row}="","",IF(WEEKDAY(B{row},2)>=6,"вых","раб"))')
        for col in (7, 8, 9, 10, 11, 12):
            ws.cell(row, col).fill = FILL_IN
            ws.cell(row, col).number_format = "0"
        ws.cell(row, 13).fill = FILL_IN
        ws.cell(row, 13).alignment = LEFT
        apply_grid(ws, row, row, 1, 13, font=FONT_N9)
        for col in (7, 8, 9, 10, 11, 12, 13):
            ws.cell(row, col).fill = FILL_IN
        # demo: 30 days of BD in January
        if n <= 30:
            ws.cell(row, 7, 1)
        # weekend highlight via formula rule later

    # conditional weekend
    ws.conditional_formatting.add(
        f"A{DAILY_FIRST}:M{DAILY_LAST}",
        FormulaRule(formula=[f'$F{DAILY_FIRST}="вых"'], fill=FILL_WEEKEND),
    )

    set_widths(
        ws,
        {
            "A": 5,
            "B": 13,
            "C": 16,
            "D": 8,
            "E": 8,
            "F": 8,
            "G": 8,
            "H": 9,
            "I": 14,
            "J": 16,
            "K": 12,
            "L": 12,
            "M": 28,
        },
    )
    print_setup(ws, landscape=True)
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"


def build_journal_660(wb: Workbook) -> None:
    ws = wb.create_sheet("Журнал_660")
    ws.sheet_properties.tabColor = "8B1E1E"
    # Official header
    ws.merge_cells("A1:M1")
    ws["A1"] = "Приложение № 7 к Порядку (п. 25)  —  приказ Министра обороны РФ от 30.10.2015 № 660"
    ws["A1"].font = FONT_SMALL
    ws["A1"].alignment = Alignment(horizontal="right", vertical="center")

    ws.merge_cells("A2:M2")
    ws["A2"] = (
        "ЖУРНАЛ УЧЕТА ВРЕМЕНИ ПРИВЛЕЧЕНИЯ ВОЕННОСЛУЖАЩИХ, ПРОХОДЯЩИХ ВОЕННУЮ СЛУЖБУ ПО КОНТРАКТУ, "
        "К ИСПОЛНЕНИЮ ОБЯЗАННОСТЕЙ ВОЕННОЙ СЛУЖБЫ В РАБОЧИЕ ДНИ СВЕРХ УСТАНОВЛЕННОЙ ПРОДОЛЖИТЕЛЬНОСТИ "
        "ЕЖЕНЕДЕЛЬНОГО СЛУЖЕБНОГО ВРЕМЕНИ, ПРИВЛЕЧЕНИЯ ЭТИХ ВОЕННОСЛУЖАЩИХ К ИСПОЛНЕНИЮ ОБЯЗАННОСТЕЙ "
        "ВОЕННОЙ СЛУЖБЫ В ВЫХОДНЫЕ И ПРАЗДНИЧНЫЕ ДНИ И ПРЕДОСТАВЛЕНИЯ ИМ ДОПОЛНИТЕЛЬНОГО ВРЕМЕНИ ОТДЫХА"
    )
    ws["A2"].font = Font(name="Times New Roman", size=10, bold=True)
    ws["A2"].alignment = CENTER
    ws.row_dimensions[2].height = 48

    ws.merge_cells("A3:B3")
    ws["A3"] = "на"
    ws["A3"].font = FONT_H
    ws["A3"].alignment = Alignment(horizontal="right", vertical="center")
    ws["C3"] = "=Анкета!C16"
    ws["C3"].fill = FILL_IN
    ws["C3"].font = FONT_H
    ws["C3"].alignment = CENTER
    ws["C3"].border = THIN
    ws["D3"] = "неделю"
    ws["D3"].font = FONT_H
    ws.merge_cells("E3:G3")
    ws["E3"] = '=IFERROR(INDEX(Журнал_день!B$6:B$371,MATCH(C3,Журнал_день!D$6:D$371,0)),"")'
    ws["E3"].number_format = "MMMM"
    ws["E3"].fill = FILL_OUT
    ws["E3"].alignment = CENTER
    ws["E3"].font = FONT_H
    ws["H3"] = "=Анкета!C12"
    ws["H3"].fill = FILL_OUT
    ws["H3"].font = FONT_H
    ws["H3"].alignment = CENTER
    ws["H3"].border = THIN
    ws["I3"] = "г."
    ws["I3"].font = FONT_H
    ws.merge_cells("J3:M3")
    ws["J3"] = '=CONCATENATE(Анкета!C8,"   в/ч ",Анкета!C9)'
    ws["J3"].font = FONT_N
    ws["J3"].alignment = CENTER

    # Multi-row official header
    # Row 4-6 headers
    heads_top = {
        1: ("№\nп/п", 4, 6, 1, 1),
        2: ("Воинская\nдолжность", 4, 6, 2, 2),
        3: ("Воинское\nзвание", 4, 6, 3, 3),
        4: ("Фамилия, имя,\nотчество", 4, 6, 4, 4),
        5: ("Учет сверхурочного\nвремени, ч", 4, 6, 5, 5),
        6: ("Учет времени привлечения\nв выходные и праздничные дни, ч", 4, 6, 6, 6),
        7: ("Суммарное время\nпривлечения к военной службе, ч", 4, 6, 7, 7),
        8: ("Учет предоставления дополнительного\nвремени отдыха", 4, 4, 8, 9),
        10: ("Учет присоединенных дополнительных\nсуток отдыха к отпуску", 4, 4, 10, 11),
        12: ("Учет нереализованного\nвремени отдыха, ч, сут.", 4, 6, 12, 12),
        13: ("Подпись\nвоеннослужащего", 4, 6, 13, 13),
    }
    for _k, (title, r1, r2, c1, c2) in heads_top.items():
        if r1 != r2 or c1 != c2:
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        cell = ws.cell(r1, c1, title)
        cell.font = FONT_H9
        cell.fill = FILL_SUB
        cell.alignment = CENTER
        cell.border = THIN
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(r, c).border = THIN
                ws.cell(r, c).fill = FILL_SUB
                ws.cell(r, c).alignment = CENTER
                ws.cell(r, c).font = FONT_H9

    ws.merge_cells("H5:H6")
    ws["H5"] = "Дата"
    ws.merge_cells("I5:I6")
    ws["I5"] = "Время, ч"
    ws.merge_cells("J5:J6")
    ws["J5"] = "Дата"
    ws.merge_cells("K5:K6")
    ws["K5"] = "Время, сут."
    for col in (8, 9, 10, 11):
        ws.cell(5, col).font = FONT_H9
        ws.cell(5, col).fill = FILL_SUB
        ws.cell(5, col).alignment = CENTER
        ws.cell(5, col).border = THIN
        ws.cell(6, col).border = THIN
        ws.cell(6, col).fill = FILL_SUB

    for col in range(1, 14):
        cell = ws.cell(7, col, col)
        cell.font = FONT_H9
        cell.fill = FILL_HEAD
        cell.font = FONT_WHITE
        cell.alignment = CENTER
        cell.border = THIN
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 18

    # First row — the person from Анкета, hours from daily for selected week
    first = 8
    ws.cell(first, 1, 1)
    ws.cell(first, 2, "=Анкета!C7")
    ws.cell(first, 3, "=Анкета!C6")
    ws.cell(first, 4, "=Анкета!B19")
    # col 5 overtime hours that week
    ws.cell(
        first,
        5,
        f'=SUMIFS(Журнал_день!I${DAILY_FIRST}:I${DAILY_LAST},Журнал_день!D${DAILY_FIRST}:D${DAILY_LAST},$C$3)',
    )
    ws.cell(
        first,
        6,
        f'=SUMIFS(Журнал_день!J${DAILY_FIRST}:J${DAILY_LAST},Журнал_день!D${DAILY_FIRST}:D${DAILY_LAST},$C$3)',
    )
    ws.cell(first, 7, f"=E{first}+F{first}")
    # rest given hours that week
    ws.cell(
        first,
        9,
        f'=SUMIFS(Журнал_день!K${DAILY_FIRST}:K${DAILY_LAST},Журнал_день!D${DAILY_FIRST}:D${DAILY_LAST},$C$3)',
    )
    ws.cell(
        first,
        11,
        f'=SUMIFS(Журнал_день!L${DAILY_FIRST}:L${DAILY_LAST},Журнал_день!D${DAILY_FIRST}:D${DAILY_LAST},$C$3)',
    )
    # unrealized: from year start through this week
    # hours remaining = (sum ot + weekend) - rest hours - leave_days * workday
    # helper hours remaining (hidden col N) — без LET, чтобы открывалось в старом Excel
    ws.cell(
        first,
        14,
        f'=SUMIFS(Журнал_день!I${DAILY_FIRST}:I${DAILY_LAST},Журнал_день!D${DAILY_FIRST}:D${DAILY_LAST},"<="&$C$3)'
        f'+SUMIFS(Журнал_день!J${DAILY_FIRST}:J${DAILY_LAST},Журнал_день!D${DAILY_FIRST}:D${DAILY_LAST},"<="&$C$3)'
        f'-SUMIFS(Журнал_день!K${DAILY_FIRST}:K${DAILY_LAST},Журнал_день!D${DAILY_FIRST}:D${DAILY_LAST},"<="&$C$3)'
        f'-SUMIFS(Журнал_день!L${DAILY_FIRST}:L${DAILY_LAST},Журнал_день!D${DAILY_FIRST}:D${DAILY_LAST},"<="&$C$3)*Анкета!C15',
    )
    ws.cell(
        first,
        12,
        f'=IF(N{first}<=0,0,IF(N{first}<Анкета!C15,N{first}&" ч",INT(N{first}/Анкета!C15)&" сут. "&MOD(N{first},Анкета!C15)&" ч"))',
    )
    ws.column_dimensions["N"].hidden = True
    ws.cell(first, 8).fill = FILL_IN
    ws.cell(first, 10).fill = FILL_IN
    ws.cell(first, 13).fill = FILL_IN
    apply_grid(ws, first, first, 1, 13)
    for col in (5, 6, 7, 9, 11, 12):
        ws.cell(first, col).fill = FILL_OUT
    ws.cell(first, 8).fill = FILL_IN
    ws.cell(first, 10).fill = FILL_IN
    ws.cell(first, 13).fill = FILL_IN
    ws.row_dimensions[first].height = 22

    # Extra rows for squad
    for n in range(2, SQUAD_ROWS + 1):
        row = first + n - 1
        ws.cell(row, 1, n)
        ws.cell(row, 7, f"=IF(OR(E{row}<>\"\",F{row}<>\"\"),N(E{row})+N(F{row}),\"\")")
        apply_grid(ws, row, row, 1, 13)
        for col in (2, 3, 4, 5, 6, 8, 9, 10, 11, 13):
            ws.cell(row, col).fill = FILL_IN
        ws.cell(row, 7).fill = FILL_OUT
        ws.cell(row, 12).fill = FILL_IN
        ws.row_dimensions[row].height = 18

    foot = first + SQUAD_ROWS + 1
    ws.merge_cells(start_row=foot, start_column=1, end_row=foot, end_column=4)
    ws.cell(foot, 1, "Командир (начальник) подразделения")
    ws.cell(foot, 1).font = FONT_H
    ws.merge_cells(start_row=foot, start_column=5, end_row=foot, end_column=8)
    ws.cell(foot, 5, "=Анкета!C8")
    ws.cell(foot, 5).fill = FILL_OUT
    ws.cell(foot, 5).alignment = CENTER
    ws.merge_cells(start_row=foot + 1, start_column=1, end_row=foot + 1, end_column=13)
    ws.cell(
        foot + 1,
        1,
        "________________________________________________________________  (воинское звание, подпись, инициал имени, фамилия)",
    )
    ws.cell(foot + 1, 1).font = FONT_N

    ws.merge_cells(start_row=foot + 3, start_column=1, end_row=foot + 6, end_column=13)
    ws.cell(
        foot + 3,
        1,
        "Порядок ведения (приказ МО РФ № 660, приложение № 7): учёт ведётся ЕЖЕНЕДЕЛЬНО. "
        "Графы 2–4 — по штатно-должностной книге. Графа 5 — сверхурочка в рабочие дни (часы). "
        "Графы 5–7 и 9 — в часах. Графа 11 — в сутках. Графа 12 — в часах, если остаток меньше "
        "продолжительности служебного дня, и в сутках, если больше. Записи доводятся до военнослужащих "
        "под роспись каждую неделю. Боевое дежурство в эту форму часами не ставьте — оно на листе «БД_сутки».",
    )
    ws.cell(foot + 3, 1).alignment = JUST
    ws.cell(foot + 3, 1).font = FONT_N9
    ws.cell(foot + 3, 1).fill = FILL_NOTE

    set_widths(
        ws,
        {
            "A": 6,
            "B": 20,
            "C": 14,
            "D": 28,
            "E": 14,
            "F": 16,
            "G": 14,
            "H": 12,
            "I": 11,
            "J": 12,
            "K": 12,
            "L": 16,
            "M": 14,
        },
    )
    print_setup(ws, landscape=True)
    ws.print_area = f"A1:M{foot + 6}"
    ws.freeze_panes = "A8"


def build_bd(wb: Workbook) -> None:
    ws = wb.create_sheet("БД_сутки")
    ws.sheet_properties.tabColor = "1F4E3D"
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "УЧЁТ БОЕВОГО ДЕЖУРСТВА И ИНЫХ МЕРОПРИЯТИЙ БЕЗ ОГРАНИЧЕНИЯ СЛУЖЕБНОГО ВРЕМЕНИ  "
        "(п. 3 ст. 11 закона № 76-ФЗ, п. 2 Перечня приказа МО РФ № 492)"
    )
    ws["A1"].font = FONT_WHITE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    ws["A2"] = (
        "Время учитывается в СУТКАХ. За каждые трое суток привлечения — двое суток отдыха "
        "(п. 5 приложения № 2 к Положению о порядке прохождения военной службы). "
        "Эти сутки можно заменить деньгами по рапорту (приказ МО РФ № 80)."
    )
    ws["A2"].font = FONT_N
    ws["A2"].fill = FILL_NOTE
    ws["A2"].alignment = LEFT
    ws.row_dimensions[2].height = 32

    labels = [
        (4, "Суток боевого дежурства за год (из Журнал_день, колонка «БД»)",
         f"=SUM(Журнал_день!G{DAILY_FIRST}:G{DAILY_LAST})"),
        (5, "Положено суток отдыха (2 за каждые 3)",
         "=INT(C4/3)*2"),
        (6, "Остаток суток БД, не вошедших в тройку (не сгорает)",
         "=MOD(C4,3)"),
        (7, "Суток суточного наряда за год",
         f"=SUM(Журнал_день!H{DAILY_FIRST}:H{DAILY_LAST})"),
        (8, "Отдых уже дан, часов (журнал 660, гр. 9)",
         f"=SUM(Журнал_день!K{DAILY_FIRST}:K{DAILY_LAST})"),
        (9, "Присоединено к отпуску, суток (гр. 11)",
         f"=SUM(Журнал_день!L{DAILY_FIRST}:L{DAILY_LAST})"),
        (10, "Суток отдыха за БД ещё не использовано",
         "=MAX(0,C5-C9)"),
        (11, "Сумма компенсации за неиспользованные сутки БД, руб.",
         "=Анкета!B24*C10"),
    ]
    ws["A3"] = "Показатель"
    ws["C3"] = "Значение"
    ws["A3"].font = FONT_WHITE
    ws["C3"].font = FONT_WHITE
    ws["A3"].fill = FILL_HEAD
    ws["C3"].fill = FILL_HEAD
    ws.merge_cells("A3:B3")
    for col in range(1, 4):
        ws.cell(3, col).border = THIN
        ws.cell(3, col).fill = FILL_HEAD
        ws.cell(3, col).alignment = CENTER

    for row, label, formula in labels:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row, 1, label).font = FONT_N
        ws.cell(row, 1).alignment = LEFT
        ws.cell(row, 1).border = THIN
        ws.cell(row, 1).fill = FILL_SUB
        ws.cell(row, 2).border = THIN
        ws.cell(row, 2).fill = FILL_SUB
        cell = ws.cell(row, 3, formula)
        cell.fill = FILL_OUT
        cell.font = FONT_H
        cell.alignment = CENTER
        cell.border = THIN
        ws.row_dimensions[row].height = 20
    ws["C11"].number_format = "#,##0.00"

    # Monthly breakdown
    ws["A13"] = "Разбивка по месяцам (нарастающий итог отдыха — с начала года, не помесячно)"
    ws.merge_cells("A13:G13")
    ws["A13"].font = FONT_WHITE
    ws["A13"].fill = FILL_HEAD
    ws["A13"].alignment = CENTER

    month_heads = [
        "Месяц",
        "Суток БД",
        "Наряд, сут.",
        "Сверхурочно, ч",
        "Выходные, ч",
        "БД с начала года",
        "Положено отдыха нарастающим",
    ]
    for col, h in enumerate(month_heads, 1):
        cell = ws.cell(14, col, h)
        cell.font = FONT_H9
        cell.fill = FILL_SUB
        cell.alignment = CENTER
        cell.border = THIN
    months = [
        "январь", "февраль", "март", "апрель", "май", "июнь",
        "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь",
    ]
    for i, name in enumerate(months, 1):
        row = 14 + i
        ws.cell(row, 1, name)
        ws.cell(row, 2, f'=SUMIF(Журнал_день!$E${DAILY_FIRST}:$E${DAILY_LAST},{i},Журнал_день!$G${DAILY_FIRST}:$G${DAILY_LAST})')
        ws.cell(row, 3, f'=SUMIF(Журнал_день!$E${DAILY_FIRST}:$E${DAILY_LAST},{i},Журнал_день!$H${DAILY_FIRST}:$H${DAILY_LAST})')
        ws.cell(row, 4, f'=SUMIF(Журнал_день!$E${DAILY_FIRST}:$E${DAILY_LAST},{i},Журнал_день!$I${DAILY_FIRST}:$I${DAILY_LAST})')
        ws.cell(row, 5, f'=SUMIF(Журнал_день!$E${DAILY_FIRST}:$E${DAILY_LAST},{i},Журнал_день!$J${DAILY_FIRST}:$J${DAILY_LAST})')
        if i == 1:
            ws.cell(row, 6, f"=B{row}")
        else:
            ws.cell(row, 6, f"=F{row-1}+B{row}")
        ws.cell(row, 7, f"=INT(F{row}/3)*2")
        apply_grid(ws, row, row, 1, 7, font=FONT_N)
        for col in range(2, 8):
            ws.cell(row, col).fill = FILL_OUT
        ws.cell(row, 1).fill = FILL_WHITE
    total_row = 27
    ws.cell(total_row, 1, "ИТОГО за год")
    ws.cell(total_row, 2, "=SUM(B15:B26)")
    ws.cell(total_row, 3, "=SUM(C15:C26)")
    ws.cell(total_row, 4, "=SUM(D15:D26)")
    ws.cell(total_row, 5, "=SUM(E15:E26)")
    ws.cell(total_row, 6, "=F26")
    ws.cell(total_row, 7, "=G26")
    apply_grid(ws, total_row, total_row, 1, 7, font=FONT_H, fill=FILL_TOTAL)

    ws["A29"] = (
        "Почему отдых не считают отдельно за январь и за февраль: остаток тройки переходит. "
        "2 суток БД в январе + 1 сутки в феврале = 2 суток отдыха, а не ноль."
    )
    ws.merge_cells("A29:G29")
    ws["A29"].font = FONT_SMALL
    ws["A29"].alignment = LEFT

    set_widths(ws, {"A": 62, "B": 14, "C": 14, "D": 16, "E": 14, "F": 20, "G": 26})
    print_setup(ws, landscape=True)


def build_totals(wb: Workbook) -> None:
    ws = wb.create_sheet("Итоги")
    ws.sheet_properties.tabColor = "C9A227"
    ws.merge_cells("A1:D1")
    ws["A1"] = "ИТОГИ ЗА ДЕНЬ, МЕСЯЦ И ГОД"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 26

    ws["A3"] = "Дата, которую смотрим (за день)"
    ws["C3"] = "=DATE(Анкета!C12,1,1)"
    ws["C3"].fill = FILL_IN
    ws["C3"].number_format = "DD.MM.YYYY"
    ws["C3"].font = FONT_H
    ws["C3"].border = THIN
    ws["A3"].font = FONT_H
    ws["A3"].border = THIN
    ws.merge_cells("A3:B3")
    ws["D3"] = "← поменяйте дату в жёлтой ячейке"
    ws["D3"].font = FONT_SMALL

    day_rows = [
        (5, "Этот день — боевое дежурство?",
         f'=IFERROR(IF(SUMIFS(Журнал_день!G${DAILY_FIRST}:G${DAILY_LAST},Журнал_день!B${DAILY_FIRST}:B${DAILY_LAST},C3)>=1,"да","нет"),"—")'),
        (6, "Этот день — суточный наряд?",
         f'=IFERROR(IF(SUMIFS(Журнал_день!H${DAILY_FIRST}:H${DAILY_LAST},Журнал_день!B${DAILY_FIRST}:B${DAILY_LAST},C3)>=1,"да","нет"),"—")'),
        (7, "Сверхурочно в этот день, ч",
         f'=IFERROR(SUMIFS(Журнал_день!I${DAILY_FIRST}:I${DAILY_LAST},Журнал_день!B${DAILY_FIRST}:B${DAILY_LAST},C3),0)'),
        (8, "Выходной/праздник в этот день, ч",
         f'=IFERROR(SUMIFS(Журнал_день!J${DAILY_FIRST}:J${DAILY_LAST},Журнал_день!B${DAILY_FIRST}:B${DAILY_LAST},C3),0)'),
        (9, "Суток БД с 1 января по эту дату",
         f'=SUMIFS(Журнал_день!G${DAILY_FIRST}:G${DAILY_LAST},Журнал_день!B${DAILY_FIRST}:B${DAILY_LAST},"<="&C3,Журнал_день!B${DAILY_FIRST}:B${DAILY_LAST},">="&DATE(Анкета!C12,1,1))'),
        (10, "Положено суток отдыха на эту дату",
         "=INT(C9/3)*2"),
        (11, "Месяц выбранной даты",
         '=CHOOSE(MONTH(C3),"январь","февраль","март","апрель","май","июнь","июль","август","сентябрь","октябрь","ноябрь","декабрь")'),
    ]
    for row, label, formula in day_rows:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row, 1, label).font = FONT_N
        ws.cell(row, 1).border = THIN
        ws.cell(row, 1).fill = FILL_SUB
        ws.cell(row, 2).border = THIN
        ws.cell(row, 2).fill = FILL_SUB
        cell = ws.cell(row, 3, formula)
        cell.fill = FILL_OUT
        cell.font = FONT_H
        cell.border = THIN
        cell.alignment = CENTER

    ws["A13"] = "За год (для рапорта)"
    ws.merge_cells("A13:C13")
    ws["A13"].font = FONT_WHITE
    ws["A13"].fill = FILL_HEAD

    year_rows = [
        (14, "Суток БД за год", "=БД_сутки!C4"),
        (15, "Положено суток отдыха за БД", "=БД_сутки!C5"),
        (16, "Уже присоединено к отпуску, сут.", "=БД_сутки!C9"),
        (17, "К выплате / неиспользовано, сут.", "=БД_сутки!C10"),
        (18, "Сумма компенсации, руб.", "=БД_сутки!C11"),
        (19, "Сверхурочно за год, ч (журнал 660)", "=БД_сутки!D27"),
        (20, "Выходные/праздники за год, ч", "=БД_сутки!E27"),
        (21, "Наряд за год, сут. (деньги по закону обычно не положены)", "=БД_сутки!C7"),
    ]
    for row, label, formula in year_rows:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row, 1, label).font = FONT_N
        ws.cell(row, 1).border = THIN
        ws.cell(row, 1).fill = FILL_SUB
        cell = ws.cell(row, 3, formula)
        cell.fill = FILL_OUT
        cell.font = FONT_H
        cell.border = THIN
        cell.alignment = CENTER
    ws["C18"].number_format = "#,##0.00"

    ws.merge_cells("A23:D25")
    ws["A23"] = (
        "Денежная компенсация в рапорте берётся из строки «К выплате». "
        "Сверхурочка и наряд из журнала 660 компенсируются отдыхом, а не деньгами "
        "(п. 1 ст. 11 закона № 76-ФЗ). Если часть отгулов уже брали — укажите их "
        "в Журнал_день в колонке «К отпуску, сут»."
    )
    ws["A23"].alignment = JUST
    ws["A23"].fill = FILL_NOTE
    ws["A23"].font = FONT_N

    set_widths(ws, {"A": 36, "B": 28, "C": 18, "D": 42})
    print_setup(ws)


def build_report(wb: Workbook) -> None:
    ws = wb.create_sheet("Рапорт")
    ws.sheet_properties.tabColor = "1F4E3D"
    ws.merge_cells("A1:F1")
    ws["A1"] = "РАПОРТ — лист готов к печати. Фамилия подставляется из анкеты."
    ws["A1"].font = FONT_WHITE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 22

    # Right-aligned addressee block
    ws.merge_cells("D3:F3")
    ws["D3"] = "=Анкета!B21"
    ws["D3"].font = FONT_N
    ws["D3"].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 32

    ws.merge_cells("D5:F5")
    ws["D5"] = "=Анкета!B22"
    ws["D5"].font = FONT_N
    ws["D5"].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    ws.row_dimensions[5].height = 32

    ws.merge_cells("A8:F8")
    ws["A8"] = "РАПОРТ"
    ws["A8"].font = Font(name="Times New Roman", size=18, bold=True)
    ws["A8"].alignment = CENTER

    ws.merge_cells("A10:F12")
    ws["A10"] = (
        '=CONCATENATE("Докладываю, что в период с 01.01.",Анкета!C12," по 31.12.",Анкета!C12,'
        '" я, ",Анкета!C6," ",Анкета!B19,", привлекался к боевому дежурству (боевой службе), '
        'проводимому без ограничения общей продолжительности еженедельного служебного времени.")'
    )
    ws["A10"].font = Font(name="Times New Roman", size=13)
    ws["A10"].alignment = JUST

    ws.merge_cells("A14:F15")
    ws["A14"] = (
        '=CONCATENATE("Количество суток привлечения к боевому дежурству — ",БД_сутки!C4,'
        '" суток. Согласно пункту 5 приложения № 2 к Положению о порядке прохождения военной службы '
        'за каждые трое суток привлечения положено двое суток отдыха.")'
    )
    ws["A14"].font = Font(name="Times New Roman", size=13)
    ws["A14"].alignment = JUST

    ws.merge_cells("A17:F18")
    ws["A17"] = (
        '=CONCATENATE("Расчёт: ",БД_сутки!C4," ÷ 3 × 2 = ",БД_сутки!C5,'
        '" дополнительных суток отдыха. Из них предоставлено (присоединено к отпуску): ",'
        'БД_сутки!C9," суток. Остаток, не использованный в виде отдыха: ",БД_сутки!C10," суток.")'
    )
    ws["A17"].font = Font(name="Times New Roman", size=13)
    ws["A17"].alignment = JUST

    ws.merge_cells("A20:F22")
    ws["A20"] = (
        '=CONCATENATE("В соответствии с пунктом 3 статьи 11 Федерального закона от 27.05.1998 № 76-ФЗ '
        '«О статусе военнослужащих» и приказом Министра обороны Российской Федерации от 14.02.2010 № 80 '
        'прошу вместо предоставления указанных дополнительных суток отдыха выплатить мне, ",'
        'Анкета!C6,"у ",Анкета!B20,", денежную компенсацию за ",БД_сутки!C10," суток.")'
    )
    ws["A20"].font = Font(name="Times New Roman", size=13)
    ws["A20"].alignment = JUST

    ws.merge_cells("A24:F25")
    ws["A24"] = (
        '=CONCATENATE("Оклад по воинской должности: ",Анкета!C13," руб.  '
        'Оклад по воинскому званию: ",Анкета!C14," руб.  '
        'Расчёт компенсации: (",Анкета!C13," + ",Анкета!C14,") ÷ 30 × ",'
        'БД_сутки!C10," = ",БД_сутки!C11," руб.")'
    )
    ws["A24"].font = Font(name="Times New Roman", size=13)
    ws["A24"].alignment = JUST

    ws.merge_cells("A27:F28")
    ws["A27"] = (
        "Дополнительные сутки отдыха за указанный период, кроме отражённых в расчёте, "
        "мне не предоставлялись. Расчёт составлен по журналу учёта служебного времени "
        "(приказ МО РФ от 30.10.2015 № 660, приложение № 7) и отметкам о боевом дежурстве."
    )
    ws["A27"].font = Font(name="Times New Roman", size=13)
    ws["A27"].alignment = JUST

    ws.merge_cells("A30:F31")
    ws["A30"] = (
        "Приложения:\n"
        "1. Выписка (копия) из журнала учёта служебного времени (приложение № 7 к приказу МО РФ № 660).\n"
        "2. Расчёт дополнительных суток отдыха за боевое дежурство."
    )
    ws["A30"].font = Font(name="Times New Roman", size=13)
    ws["A30"].alignment = LEFT

    ws.merge_cells("A34:B34")
    ws["A34"] = '=CONCATENATE("«    » ______________ ",Анкета!C12," г.")'
    ws["A34"].font = Font(name="Times New Roman", size=13)

    ws.merge_cells("D34:F34")
    ws["D34"] = '=CONCATENATE("__________ / ",Анкета!C6," ",Анкета!B19," /")'
    ws["D34"].font = Font(name="Times New Roman", size=13)
    ws["D34"].alignment = Alignment(horizontal="right", vertical="center")

    ws.merge_cells("A37:F38")
    ws["A37"] = (
        "Служебная отметка канцелярии: входящий № ________ от «    » ______________ 20    г."
    )
    ws["A37"].font = FONT_N
    ws["A37"].fill = FILL_NOTE
    ws["A37"].alignment = LEFT

    set_widths(ws, {"A": 16, "B": 16, "C": 16, "D": 18, "E": 18, "F": 22})
    for r in (10, 14, 17, 20, 24, 27, 30):
        ws.row_dimensions[r].height = 22
    print_setup(ws, landscape=False)
    ws.print_area = "A1:F38"
    ws.page_setup.fitToHeight = 1


def main() -> None:
    wb = Workbook()
    build_help(wb)
    build_anketa(wb)
    build_daily(wb)
    build_journal_660(wb)
    build_bd(wb)
    build_totals(wb)
    build_report(wb)

    # named ranges for robustness
    from openpyxl.workbook.defined_name import DefinedName

    wb.defined_names.add(DefinedName(name="Год", attr_text="Анкета!$C$12"))
    wb.defined_names.add(DefinedName(name="Фамилия", attr_text="Анкета!$C$3"))
    wb.defined_names.add(DefinedName(name="ОВД", attr_text="Анкета!$C$13"))
    wb.defined_names.add(DefinedName(name="ОВЗ", attr_text="Анкета!$C$14"))

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.oddFooter.center.text = "&K000000 Журнал и рапорт — приказ МО РФ № 660 прил. 7, приказ МО РФ № 80"

    wb.save(OUT)
    print(f"saved {OUT}")


if __name__ == "__main__":
    main()
